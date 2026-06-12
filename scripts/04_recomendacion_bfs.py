#%%
# ============================================================
# Objetivo:
# Implementar una recomendación de películas usando la lógica de
# BFS sobre un grafo bipartito Usuario-Película.
#
# El recorrido se interpreta así:
#
# Nivel 0: Usuario objetivo
# Nivel 1: Películas calificadas positivamente por el usuario
# Nivel 2: Usuarios similares que calificaron esas mismas películas
# Nivel 3: Películas candidatas calificadas positivamente por usuarios similares
# ============================================================

# ============================================================
# 1. Importar librerías
# ============================================================
# Descargar desde CMD con el comando: pip install openpyxl , para exportar a Excel con formato.
import pandas as pd
import networkx as nx
from collections import deque, defaultdict
from pyvis.network import Network

try:
    from IPython.display import IFrame
    USAR_IFRAME = True
except ImportError:
    USAR_IFRAME = False

#%%
# ============================================================
# 2. Cargar dataset y grafo bipartito
# ============================================================

# Dataset preprocesado
df_2000 = pd.read_csv("data/peliculas_2000.csv")

# Grafo construido en el archivo 03_grafo_bipartito_con_networkx.py
G = nx.read_gml("data/grafo_bipartito_peliculas.gml")

print("Dataset y grafo cargados correctamente.")
print("Cantidad de registros:", len(df_2000))
print("Cantidad de nodos:", G.number_of_nodes())
print("Cantidad de aristas:", G.number_of_edges())


#%%
# ============================================================
# 3. Separar usuarios y películas del grafo
# ============================================================

usuarios_grafo = [
    n for n, d in G.nodes(data=True)
    if d["tipo"] == "usuario"
]

peliculas_grafo = [
    n for n, d in G.nodes(data=True)
    if d["tipo"] == "pelicula"
]

print("Usuarios en el grafo:", len(usuarios_grafo))
print("Películas en el grafo:", len(peliculas_grafo))


#%%
# ============================================================
# 4. Seleccionar usuario objetivo
# ============================================================

# Para la prueba inicial, se elige el usuario con más conexiones.
# Esto permite obtener más información para generar recomendaciones.

usuario_objetivo = max(usuarios_grafo, key=lambda u: G.degree(u))

print("Usuario objetivo seleccionado:", usuario_objetivo)
print("Cantidad de películas calificadas:", G.degree(usuario_objetivo))

#%%
# ============================================================
# 5. Función auxiliar: obtener rating de una arista
# ============================================================

def obtener_rating(G, nodo1, nodo2):
    """
    Devuelve el rating almacenado en la arista entre nodo1 y nodo2.
    Si no existe rating, retorna 0.
    """
    return G[nodo1][nodo2].get("rating", 0)


#%%
# ============================================================
# 6. Recomendación usando lógica BFS
# ============================================================

def recomendar_peliculas_bfs(G, usuario_objetivo, rating_min=4.0, top_n=10):
    """
    Recomienda películas usando una lógica basada en BFS
    sobre un grafo bipartito Usuario-Película.

    Parámetros:
    - G: grafo bipartito construido con NetworkX.
    - usuario_objetivo: nodo del usuario objetivo, por ejemplo 'U_1'.
    - rating_min: calificación mínima considerada positiva.
    - top_n: cantidad de recomendaciones a devolver.

    Retorna:
    - DataFrame con las películas recomendadas.
    """

    # ------------------------------------------------------------
    # Nivel 0: usuario objetivo
    # ------------------------------------------------------------

    if usuario_objetivo not in G:
        print("El usuario objetivo no existe en el grafo.")
        return pd.DataFrame()

    # ------------------------------------------------------------
    # Nivel 1: películas calificadas positivamente por el usuario
    # ------------------------------------------------------------

    peliculas_usuario = set()
    peliculas_positivas_usuario = set()

    for pelicula in G.neighbors(usuario_objetivo):
        if G.nodes[pelicula]["tipo"] == "pelicula":
            peliculas_usuario.add(pelicula)

            rating = obtener_rating(G, usuario_objetivo, pelicula)

            if rating >= rating_min:
                peliculas_positivas_usuario.add(pelicula)

    print("Películas calificadas por el usuario:", len(peliculas_usuario))
    print("Películas positivas del usuario:", len(peliculas_positivas_usuario))

    if len(peliculas_positivas_usuario) == 0:
        print("El usuario no tiene suficientes películas con calificación positiva.")
        return pd.DataFrame()

    # ------------------------------------------------------------
    # Nivel 2: usuarios similares
    # ------------------------------------------------------------

    usuarios_similares = set()

    for pelicula in peliculas_positivas_usuario:
        for usuario in G.neighbors(pelicula):
            if (
                usuario != usuario_objetivo
                and G.nodes[usuario]["tipo"] == "usuario"
            ):
                rating = obtener_rating(G, usuario, pelicula)

                if rating >= rating_min:
                    usuarios_similares.add(usuario)

    print("Usuarios similares encontrados:", len(usuarios_similares))

    if len(usuarios_similares) == 0:
        print("No se encontraron usuarios similares.")
        return pd.DataFrame()

    # ------------------------------------------------------------
    # Nivel 3: películas candidatas
    # ------------------------------------------------------------

    peliculas_candidatas = defaultdict(list)

    for usuario in usuarios_similares:
        for pelicula in G.neighbors(usuario):
            if G.nodes[pelicula]["tipo"] == "pelicula":

                # No recomendar películas que el usuario objetivo ya calificó
                if pelicula in peliculas_usuario:
                    continue

                rating = obtener_rating(G, usuario, pelicula)

                if rating >= rating_min:
                    peliculas_candidatas[pelicula].append(rating)

    print("Películas candidatas encontradas:", len(peliculas_candidatas))

    if len(peliculas_candidatas) == 0:
        print("No se encontraron películas candidatas.")
        return pd.DataFrame()

    # ------------------------------------------------------------
    # Ranking de recomendaciones
    # ------------------------------------------------------------

    recomendaciones = []

    for pelicula, ratings in peliculas_candidatas.items():
        datos_pelicula = G.nodes[pelicula]

        rating_promedio = sum(ratings) / len(ratings)
        cantidad_votos = len(ratings)

        recomendaciones.append({
            "movie_node": pelicula,
            "titulo": datos_pelicula.get("titulo", pelicula),
            "year": datos_pelicula.get("year", ""),
            "generos": datos_pelicula.get("generos", ""),
            "rating_promedio": round(rating_promedio, 2),
            "cantidad_votos_similares": cantidad_votos
        })

    recomendaciones_df = pd.DataFrame(recomendaciones)

    recomendaciones_df = recomendaciones_df.sort_values(
        by=["rating_promedio", "cantidad_votos_similares"],
        ascending=False
    )

    return recomendaciones_df.head(top_n)


#%%
# ============================================================
# 7. Ejecutar recomendación
# ============================================================

recomendaciones = recomendar_peliculas_bfs(
    G,
    usuario_objetivo=usuario_objetivo,
    rating_min=4.0,
    top_n=10
)

print("\n========== RECOMENDACIONES ==========")
print(recomendaciones)
print("=====================================\n")

#%%
# ============================================================
# 8. Visualización interactiva del proceso de recomendación
# ============================================================

def crear_grafo_recomendacion_pyvis(
    G,
    usuario_objetivo,
    recomendaciones_df,
    rating_min=4.0,
    max_peliculas_usuario=8,
    max_peliculas_no_recomendadas_por_usuario=2,
    archivo_html="grafo_recomendacion_bfs.html"
):
    """
    Crea una visualización interactiva conectada del proceso de recomendación BFS.

    Esta versión evita islas:
    - No agrega usuarios solos.
    - No agrega películas solas.
    - Todas las recomendaciones rojas se intentan conectar mediante un camino real:

      Usuario objetivo -> Película positiva en común -> Usuario similar -> Película recomendada

    Colores:
    - Usuario objetivo: naranja
    - Películas positivas del usuario objetivo: verde
    - Usuarios similares: celeste
    - Películas recomendadas: rojo
    - Películas no recomendadas: gris oscuro
    """

    # ------------------------------------------------------------
    # 8.1. Obtener películas del usuario objetivo
    # ------------------------------------------------------------

    peliculas_usuario = {
        p for p in G.neighbors(usuario_objetivo)
        if G.nodes[p]["tipo"] == "pelicula"
    }

    peliculas_positivas_usuario = []

    for pelicula in peliculas_usuario:
        rating = obtener_rating(G, usuario_objetivo, pelicula)

        if rating >= rating_min:
            peliculas_positivas_usuario.append(pelicula)

    # ------------------------------------------------------------
    # 8.2. Obtener películas recomendadas desde recomendaciones_df
    # ------------------------------------------------------------

    peliculas_recomendadas = []

    if recomendaciones_df is not None and not recomendaciones_df.empty:
        peliculas_recomendadas = list(recomendaciones_df["movie_node"].astype(str))

    print("Películas recomendadas que se mostrarán en rojo:", len(peliculas_recomendadas))

    # ------------------------------------------------------------
    # 8.3. Crear red interactiva
    # ------------------------------------------------------------

    net = Network(
        notebook=True,
        cdn_resources="in_line",
        height="800px",
        width="100%",
        bgcolor="#ffffff",
        font_color="black"
    )

    # ------------------------------------------------------------
    # 8.4. Funciones auxiliares
    # ------------------------------------------------------------

    nodos_agregados = set()
    aristas_agregadas = set()

    def agregar_arista(u, v, label, title, color):
        clave = tuple(sorted([u, v]))

        if clave not in aristas_agregadas:
            net.add_edge(
                u,
                v,
                label=str(label),
                title=title,
                color=color
            )
            aristas_agregadas.add(clave)

    def agregar_usuario_objetivo():
        if usuario_objetivo not in nodos_agregados:
            net.add_node(
                usuario_objetivo,
                label=usuario_objetivo,
                title="Usuario objetivo",
                color="orange",
                size=34
            )
            nodos_agregados.add(usuario_objetivo)

    def agregar_pelicula_positiva(pelicula):
        if pelicula not in nodos_agregados:
            datos = G.nodes[pelicula]
            titulo = datos.get("titulo", pelicula)
            year = datos.get("year", "")
            generos = datos.get("generos", "")
            rating = obtener_rating(G, usuario_objetivo, pelicula)

            net.add_node(
                pelicula,
                label=titulo,
                title=(
                    f"Película positiva del usuario objetivo\n"
                    f"Título: {titulo}\n"
                    f"Año: {year}\n"
                    f"Géneros: {generos}\n"
                    f"Rating del usuario objetivo: {rating}"
                ),
                color="lightgreen",
                size=22
            )
            nodos_agregados.add(pelicula)

        rating_objetivo = obtener_rating(G, usuario_objetivo, pelicula)

        agregar_arista(
            usuario_objetivo,
            pelicula,
            rating_objetivo,
            f"Rating del usuario objetivo: {rating_objetivo}",
            "green"
        )

    def agregar_usuario_similar(usuario):
        if usuario not in nodos_agregados:
            net.add_node(
                usuario,
                label=usuario,
                title="Usuario similar",
                color="lightblue",
                size=18
            )
            nodos_agregados.add(usuario)

    def agregar_pelicula_recomendada(pelicula):
        if pelicula not in nodos_agregados:
            datos = G.nodes[pelicula]
            titulo = datos.get("titulo", pelicula)
            year = datos.get("year", "")
            generos = datos.get("generos", "")

            fila_rec = recomendaciones_df[
                recomendaciones_df["movie_node"].astype(str) == pelicula
            ]

            if not fila_rec.empty:
                rating_promedio = fila_rec.iloc[0].get("rating_promedio", "")
                votos = fila_rec.iloc[0].get("cantidad_votos_similares", "")
            else:
                rating_promedio = ""
                votos = ""

            net.add_node(
                pelicula,
                label=titulo,
                title=(
                    f"PELÍCULA RECOMENDADA\n"
                    f"Título: {titulo}\n"
                    f"Año: {year}\n"
                    f"Géneros: {generos}\n"
                    f"Rating promedio: {rating_promedio}\n"
                    f"Votos de usuarios similares: {votos}"
                ),
                color="red",
                size=30
            )
            nodos_agregados.add(pelicula)

    def agregar_pelicula_no_recomendada(pelicula):
        if pelicula not in nodos_agregados:
            datos = G.nodes[pelicula]
            titulo = datos.get("titulo", pelicula)
            year = datos.get("year", "")
            generos = datos.get("generos", "")

            net.add_node(
                pelicula,
                label=titulo,
                title=(
                    f"Película no recomendada\n"
                    f"Título: {titulo}\n"
                    f"Año: {year}\n"
                    f"Géneros: {generos}\n"
                    f"Motivo: rating menor a {rating_min}"
                ),
                color="darkgray",
                size=15
            )
            nodos_agregados.add(pelicula)

    # ------------------------------------------------------------
    # 8.5. Agregar usuario objetivo
    # ------------------------------------------------------------

    agregar_usuario_objetivo()

    # ------------------------------------------------------------
    # 8.6. Construir caminos conectados para todas las recomendaciones
    # ------------------------------------------------------------

    usuarios_conectados = set()
    peliculas_puente_conectadas = set()

    for pelicula_recomendada in peliculas_recomendadas:

        if pelicula_recomendada not in G:
            continue

        camino_encontrado = False

        # Buscar un usuario que calificó positivamente la película recomendada
        for usuario_similar in G.neighbors(pelicula_recomendada):

            if G.nodes[usuario_similar]["tipo"] != "usuario":
                continue

            rating_recomendada = obtener_rating(G, usuario_similar, pelicula_recomendada)

            if rating_recomendada < rating_min:
                continue

            # Buscar una película puente:
            # debe estar calificada positivamente por el usuario objetivo
            # y también positivamente por el usuario similar.
            for pelicula_puente in peliculas_positivas_usuario:

                if not G.has_edge(usuario_similar, pelicula_puente):
                    continue

                rating_objetivo = obtener_rating(G, usuario_objetivo, pelicula_puente)
                rating_similar_puente = obtener_rating(G, usuario_similar, pelicula_puente)

                if rating_objetivo >= rating_min and rating_similar_puente >= rating_min:

                    # Agregar camino completo:
                    # Usuario objetivo -> Película puente -> Usuario similar -> Película recomendada
                    agregar_pelicula_positiva(pelicula_puente)
                    agregar_usuario_similar(usuario_similar)
                    agregar_pelicula_recomendada(pelicula_recomendada)

                    agregar_arista(
                        pelicula_puente,
                        usuario_similar,
                        rating_similar_puente,
                        f"Rating del usuario similar: {rating_similar_puente}",
                        "blue"
                    )

                    agregar_arista(
                        usuario_similar,
                        pelicula_recomendada,
                        rating_recomendada,
                        f"Rating positivo de {usuario_similar}: {rating_recomendada}",
                        "red"
                    )

                    usuarios_conectados.add(usuario_similar)
                    peliculas_puente_conectadas.add(pelicula_puente)

                    camino_encontrado = True
                    break

            if camino_encontrado:
                break

        if not camino_encontrado:
            print(
                f"Advertencia: no se pudo conectar visualmente la recomendación {pelicula_recomendada}"
            )

    # ------------------------------------------------------------
    # 8.7. Agregar algunas películas positivas adicionales del usuario objetivo
    #      solo conectadas al usuario objetivo, para contexto visual
    # ------------------------------------------------------------

    contador_contexto = 0

    for pelicula in peliculas_positivas_usuario:

        if contador_contexto >= max_peliculas_usuario:
            break

        if pelicula not in peliculas_puente_conectadas:
            agregar_pelicula_positiva(pelicula)
            contador_contexto += 1

    # ------------------------------------------------------------
    # 8.8. Agregar películas no recomendadas en gris oscuro
    #      solo desde usuarios ya conectados
    # ------------------------------------------------------------

    peliculas_recomendadas_set = set(peliculas_recomendadas)

    for usuario in usuarios_conectados:
        contador_grises = 0

        for pelicula in G.neighbors(usuario):

            if G.nodes[pelicula]["tipo"] != "pelicula":
                continue

            # No mostrar películas ya calificadas por el usuario objetivo
            if pelicula in peliculas_usuario:
                continue

            # No mostrar recomendaciones como grises
            if pelicula in peliculas_recomendadas_set:
                continue

            rating = obtener_rating(G, usuario, pelicula)

            # Solo se muestran películas no recomendadas con rating bajo
            if rating >= rating_min:
                continue

            if contador_grises >= max_peliculas_no_recomendadas_por_usuario:
                break

            agregar_pelicula_no_recomendada(pelicula)

            agregar_arista(
                usuario,
                pelicula,
                rating,
                f"Rating otorgado por {usuario}: {rating}",
                "darkgray"
            )

            contador_grises += 1

    # ------------------------------------------------------------
    # 8.9. Guardar HTML
    # ------------------------------------------------------------

    net.toggle_physics(True)

    html_content = net.generate_html()

    with open(archivo_html, mode="w", encoding="utf-8") as f:
        f.write(html_content)

    print(f"Grafo interactivo de recomendación generado: {archivo_html}")

    if USAR_IFRAME:
        return IFrame(archivo_html, width=1000, height=800)

crear_grafo_recomendacion_pyvis(
    G,
    usuario_objetivo=usuario_objetivo,
    recomendaciones_df=recomendaciones,
    rating_min=4.0,
    max_peliculas_usuario=8,
    max_peliculas_no_recomendadas_por_usuario=2,
    archivo_html="grafo_recomendacion_bfs.html"
)

#%%
# ============================================================
# 9. Mostrar explicación del recorrido BFS aplicado
# ============================================================

print("Interpretación del recorrido BFS aplicado:")
print("Nivel 0: Usuario objetivo:", usuario_objetivo)
print("Nivel 1: Películas calificadas positivamente por el usuario.")
print("Nivel 2: Usuarios similares que calificaron positivamente esas películas.")
print("Nivel 3: Películas candidatas calificadas positivamente por usuarios similares.")


#%%
# ============================================================
# 10. Implementación explícita de BFS hasta profundidad 3
# ============================================================

def bfs_limitado(G, inicio, profundidad_max=3):
    """
    Realiza BFS desde un nodo inicial hasta una profundidad máxima.

    Retorna:
    - visitados: conjunto de nodos visitados
    - niveles: diccionario con el nivel de cada nodo
    - padres: diccionario con el padre de cada nodo en el recorrido
    """

    visitados = set()
    niveles = {}
    padres = {}

    cola = deque()

    visitados.add(inicio)
    niveles[inicio] = 0
    padres[inicio] = None
    cola.append(inicio)

    while cola:
        nodo_actual = cola.popleft()
        nivel_actual = niveles[nodo_actual]

        if nivel_actual == profundidad_max:
            continue

        for vecino in G.neighbors(nodo_actual):
            if vecino not in visitados:
                visitados.add(vecino)
                niveles[vecino] = nivel_actual + 1
                padres[vecino] = nodo_actual
                cola.append(vecino)

    return visitados, niveles, padres


visitados, niveles, padres = bfs_limitado(
    G,
    inicio=usuario_objetivo,
    profundidad_max=3
)

print("\n========== BFS LIMITADO ==========")
print("Nodos visitados hasta profundidad 3:", len(visitados))

niveles_conteo = defaultdict(int)

for nodo, nivel in niveles.items():
    niveles_conteo[nivel] += 1

for nivel in sorted(niveles_conteo.keys()):
    print(f"Nivel {nivel}: {niveles_conteo[nivel]} nodos")

print("=================================\n")


#%%
# ============================================================
# 11. Guardar recomendaciones en CSV
# ============================================================

recomendaciones.to_csv("data/recomendaciones_bfs.csv", index=False)

print("Recomendaciones guardadas en: data/recomendaciones_bfs.csv")
# %%
# ============================================================
# Exportar recomendaciones a Excel con colores por columna
# ============================================================

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

archivo_excel = "data/recomendaciones_bfs.xlsx"

columnas_exportar = [
    "movie_node",
    "titulo",
    "year",
    "generos",
    "rating_promedio",
    "cantidad_votos_similares"
]

with pd.ExcelWriter(archivo_excel, engine="openpyxl") as writer:
    recomendaciones[columnas_exportar].to_excel(
        writer,
        sheet_name="Recomendaciones",
        index=False
    )

    workbook = writer.book
    worksheet = writer.sheets["Recomendaciones"]

    # Colores de encabezado por columna
    colores_encabezado = {
        "A": "404040",  # gris oscuro para movie_node
        "B": "1F4E78",  # azul
        "C": "70AD47",  # verde
        "D": "7030A0",  # morado
        "E": "C00000",  # rojo
        "F": "ED7D31",  # naranja
    }

    # Colores suaves para celdas por columna
    colores_celdas = {
        "A": "D9D9D9",  # gris claro
        "B": "D9EAF7",  # azul claro
        "C": "E2F0D9",  # verde claro
        "D": "EADCF8",  # morado claro
        "E": "F4CCCC",  # rojo claro
        "F": "FCE4D6",  # naranja claro
    }

    borde_fino = Side(style="thin", color="BFBFBF")
    borde = Border(
        left=borde_fino,
        right=borde_fino,
        top=borde_fino,
        bottom=borde_fino
    )

    # Formato de encabezados
    for col in range(1, len(columnas_exportar) + 1):
        letra_col = get_column_letter(col)
        celda = worksheet[f"{letra_col}1"]

        celda.fill = PatternFill(
            start_color=colores_encabezado[letra_col],
            end_color=colores_encabezado[letra_col],
            fill_type="solid"
        )
        celda.font = Font(color="FFFFFF", bold=True)
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = borde

    # Formato de celdas por columna
    for row in range(2, worksheet.max_row + 1):
        for col in range(1, len(columnas_exportar) + 1):
            letra_col = get_column_letter(col)
            celda = worksheet[f"{letra_col}{row}"]

            celda.fill = PatternFill(
                start_color=colores_celdas[letra_col],
                end_color=colores_celdas[letra_col],
                fill_type="solid"
            )
            celda.border = borde
            celda.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    # Formato numérico
    for row in range(2, worksheet.max_row + 1):
        worksheet[f"E{row}"].number_format = "0.00"  # rating_promedio
        worksheet[f"F{row}"].number_format = "0"     # cantidad_votos_similares

    # Ajustar ancho de columnas
    worksheet.column_dimensions["A"].width = 20  # movie_node
    worksheet.column_dimensions["B"].width = 45  # título
    worksheet.column_dimensions["C"].width = 12  # año
    worksheet.column_dimensions["D"].width = 35  # géneros
    worksheet.column_dimensions["E"].width = 20  # rating promedio
    worksheet.column_dimensions["F"].width = 30  # cantidad votos

    # Altura de filas
    worksheet.row_dimensions[1].height = 25

    for row in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 35

    # Congelar encabezado
    worksheet.freeze_panes = "A2"

    # Agregar filtro
    worksheet.auto_filter.ref = worksheet.dimensions

print(f"Archivo Excel generado con colores correctamente: {archivo_excel}")
# %%
